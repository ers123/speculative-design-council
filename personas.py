"""Speculative Design Council — Nemotron-Personas-Korea 페르소나 통합 모듈.

기능:
  1. Nemotron-Personas-Korea Excel/JSON 로드
  2. 학생 narratype 기반 4명 위원회 자동 추천 (Ollama)
  3. 각 페르소나를 Council 역할(거울/지도/의장/미래학자)에 매핑
  4. system prompt에 페르소나 정체성 주입 (각 에이전트가 그 페르소나로 발언)
  5. Council 구성 요약 (export 및 UI 표시용)

설계 원칙:
- Backward compatible: 페르소나 없으면 기존 추상 Council 동작
- 페르소나는 prompt에 주입되는 텍스트일 뿐 — Harness 코드 변경 없음
- "같은 코드, 다른 프롬프트" 메시지 강화 (페르소나도 그 다른 프롬프트의 일부)
"""

from __future__ import annotations

import io
import json
import re
from typing import Optional

import ollama


# ═════════════════════════════════════════════════════════════════════
# 페르소나 로드
# ═════════════════════════════════════════════════════════════════════

# 우리가 기대하는 Nemotron 샘플 Excel 컬럼 (한국어 헤더)
EXPECTED_FIELDS = [
    "이름", "성별", "나이", "시도", "시군구",
    "학력", "직업", "한줄_persona",
    "직업_persona", "가족_persona", "취미", "문화배경",
]


def load_personas_from_xlsx(file_buf) -> list[dict]:
    """Read Nemotron Excel (xlsx) — returns list of persona dicts.

    file_buf: file path, BytesIO, or Streamlit UploadedFile.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.load_workbook(file_buf, data_only=True)
    # Find the personas sheet (skip README)
    sheet = None
    for name in wb.sheetnames:
        if name.lower().startswith("readme"):
            continue
        sheet = wb[name]
        break
    if sheet is None:
        sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    personas = []
    for row in rows[1:]:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        d = {h: (v if v is not None else "") for h, v in zip(headers, row)}
        personas.append(d)
    return personas


def load_personas_from_json(file_buf) -> list[dict]:
    """Read JSON list of persona dicts."""
    if hasattr(file_buf, "read"):
        data = json.load(file_buf)
    elif isinstance(file_buf, (str, bytes)):
        with open(file_buf) as f:
            data = json.load(f)
    else:
        data = file_buf
    if not isinstance(data, list):
        raise ValueError("JSON must be a list of persona dicts")
    return data


def load_personas(file_path_or_buf, ext: str = "") -> list[dict]:
    """Auto-detect xlsx vs json and load."""
    if hasattr(file_path_or_buf, "name"):
        name = file_path_or_buf.name
    else:
        name = str(file_path_or_buf)
    name_lower = (name or "").lower()
    if ext.lower() == "json" or name_lower.endswith(".json"):
        return load_personas_from_json(file_path_or_buf)
    return load_personas_from_xlsx(file_path_or_buf)


# ═════════════════════════════════════════════════════════════════════
# 페르소나 표현 (prompt 주입용)
# ═════════════════════════════════════════════════════════════════════

def format_persona_brief(persona: dict, max_chars: int = 500) -> str:
    """Compact persona description for prompt injection.

    Returns a paragraph capturing identity + voice texture.
    """
    name = persona.get("이름", "?")
    age = persona.get("나이", "?")
    sex = persona.get("성별", "?")
    region = " ".join(filter(None, [str(persona.get("시도", "")), str(persona.get("시군구", ""))]))
    occupation = persona.get("직업", "")
    education = persona.get("학력", "")
    one_liner = persona.get("한줄_persona", "")
    job_persona = persona.get("직업_persona", "")
    cultural = persona.get("문화배경", "")
    hobbies = persona.get("취미", "")

    parts = [
        f"이름: {name} ({age}세 {sex})",
        f"지역: {region}",
        f"직업·학력: {occupation} / {education}" if education else f"직업: {occupation}",
    ]
    if one_liner:
        parts.append(f"요약: {one_liner}")
    if cultural:
        parts.append(f"문화배경: {cultural[:200]}")
    if job_persona:
        parts.append(f"일터: {job_persona[:150]}")
    if hobbies:
        parts.append(f"취미·관심: {hobbies[:150]}")

    text = "\n".join(parts)
    if len(text) > max_chars:
        text = text[:max_chars] + "..."
    return text


def display_label(persona: dict) -> str:
    """One-line label for UI display."""
    return f"{persona.get('이름', '?')} ({persona.get('나이', '?')}세 {persona.get('성별', '?')}, {persona.get('시도', '?')} · {persona.get('직업', '?')[:18]})"


# ═════════════════════════════════════════════════════════════════════
# 위원회 구성 추천 (Ollama)
# ═════════════════════════════════════════════════════════════════════

ROLE_DESCRIPTIONS = {
    "scout": ("거울 (Mirror)", "narratype을 4 archetype lens(성장/붕괴/지속/변혁)로 확장하는 생성자. 다양성과 상상력 풍부한 페르소나가 좋음."),
    "critic": ("지도 (Map)", "거울의 확장에 비판·도전을 가하는 분석가. 비판적·구조적·정치적 시각이 강한 페르소나가 좋음."),
    "director": ("의장 (Chairman)", "양측을 종합해 결정을 내리는 중재자. 균형감각·중립성·결정력 있는 페르소나가 좋음."),
    "advisor": ("미래학자 (Futurist)", "메타 질문 10개를 던지는 호기심꾼. Dator 4 archetypes lens로 묻기. 상상력·세대 감각 강한 페르소나가 좋음."),
}


_RECOMMEND_PROMPT = """당신은 Speculative Design 워크숍의 Council Composer입니다.
학생의 narratype과 페르소나 후보 풀을 받아, 의미 있는 위원회 4명(거울·지도·의장·미래학자)을 추천합니다.

위원회 역할:
- 거울 (Mirror): narratype 확장 · 다양성·상상력 풍부한 페르소나
- 지도 (Map): 비평·도전 · 비판적·구조적·정치적 시각
- 의장 (Chairman): 종합·결정 · 균형감각·중립
- 미래학자 (Futurist): 메타 질문 · 호기심·세대감각

★★★ 가장 중요한 규칙 ★★★
- demographic 다양성을 적극 추구하세요 (연령·성별·지역·계층 분산)
- "서울 20대 디자인 전공"만 모이는 위원회 금지
- narratype의 사각지대를 메우는 페르소나 선택

★★★ 출력 형식 (반드시 정확히 이 형식) ★★★
거울: 인덱스 [N] | [한 줄 이유]
지도: 인덱스 [M] | [한 줄 이유]
의장: 인덱스 [P] | [한 줄 이유]
미래학자: 인덱스 [Q] | [한 줄 이유]

(N, M, P, Q는 0부터 시작하는 정수 인덱스. 같은 인덱스 중복 금지.)"""


def _format_candidate_list(personas: list[dict], max_chars_each: int = 200) -> str:
    """Build a compact numbered list of candidates for the recommender."""
    lines = []
    for i, p in enumerate(personas):
        brief = format_persona_brief(p, max_chars=max_chars_each).replace("\n", " | ")
        lines.append(f"[{i}] {brief}")
    return "\n".join(lines)


def _parse_recommendation(text: str) -> dict[str, int]:
    """Parse '거울: 인덱스 [N] | reason' style output into role -> index map."""
    role_map = {"거울": "scout", "지도": "map_critic", "의장": "director", "미래학자": "advisor"}
    # We use 'critic' internally (not 'map_critic'); align below
    role_internal = {"거울": "scout", "지도": "critic", "의장": "director", "미래학자": "advisor"}
    result: dict[str, int] = {}
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        m = re.match(r"^[\-\*\s]*([가-힣]{2,5})\s*[:：]\s*(?:인덱스\s*)?\[?\s*(\d+)\s*\]?", line)
        if not m:
            continue
        role_kor = m.group(1)
        idx = int(m.group(2))
        if role_kor in role_internal:
            result[role_internal[role_kor]] = idx
    return result


def recommend_council(
    narratype_text: str,
    personas: list[dict],
    model: str,
    n_candidates: int = 30,
) -> dict[str, int]:
    """Use Ollama to recommend 4 personas for Council roles.

    Returns: {"scout": idx, "critic": idx, "director": idx, "advisor": idx}
    Falls back to first 4 personas if recommendation fails.
    """
    if len(personas) < 4:
        # Not enough candidates; assign first 4 (or repeat if fewer)
        roles = ["scout", "critic", "director", "advisor"]
        return {r: i % max(1, len(personas)) for i, r in enumerate(roles)}

    candidates = personas[:n_candidates]
    candidate_list = _format_candidate_list(candidates)

    user_payload = (
        f"=== 학생 narratype ===\n{narratype_text[:1500]}\n=== 끝 ===\n\n"
        f"=== 페르소나 후보 {len(candidates)}명 ===\n{candidate_list}\n=== 끝 ===\n\n"
        f"위 후보 중에서 4명을 골라 출력 형식대로만 답하세요. 인덱스는 반드시 위 후보 범위(0~{len(candidates)-1}) 안에서."
    )

    try:
        response = ollama.chat(
            model=model,
            messages=[
                {"role": "system", "content": _RECOMMEND_PROMPT},
                {"role": "user", "content": user_payload},
            ],
            options={"temperature": 0.4, "num_predict": 500},
            think=False,
            keep_alive=0,
            stream=False,
        )
        msg = response.get("message") if isinstance(response, dict) else getattr(response, "message", None)
        text = (msg.get("content", "") if isinstance(msg, dict) else getattr(msg, "content", "")) or ""
    except Exception:
        text = ""

    parsed = _parse_recommendation(text)

    # Fill missing roles with sensible defaults (distinct indices)
    used = set(parsed.values())
    fallback_seq = [i for i in range(len(candidates)) if i not in used]
    for r in ["scout", "critic", "director", "advisor"]:
        if r not in parsed:
            if fallback_seq:
                parsed[r] = fallback_seq.pop(0)
            else:
                parsed[r] = 0

    # Validate index range
    for r, idx in list(parsed.items()):
        if not (0 <= idx < len(personas)):
            parsed[r] = 0

    return parsed


# ═════════════════════════════════════════════════════════════════════
# Council 구성 → prompt 주입
# ═════════════════════════════════════════════════════════════════════

def make_personified_prompt(role: str, base_prompt: str, persona: Optional[dict]) -> str:
    """Wrap a base role prompt with persona identity.

    If persona is None, returns base_prompt unchanged (backward compatible).
    """
    if persona is None:
        return base_prompt

    label, role_desc = ROLE_DESCRIPTIONS.get(role, (role, ""))
    brief = format_persona_brief(persona)

    persona_prefix = f"""
═══════════════════════════════════════════
[당신이 연기할 정체]

{brief}

[Council 역할]
{label} — {role_desc}

★★★ 페르소나 발화 규칙 ★★★
- 위 페르소나의 1인칭 목소리·어휘·관심사·우려를 그대로 살려라
- 응답은 "{persona.get('이름', '?')}입니다." 같은 1인칭 자기소개로 시작하라
- 그 페르소나가 모를 만한 전문 용어는 회피하거나 자기 식으로 풀어 표현
- 그 페르소나의 demographic이 가져오는 시각·편견·관심사를 정직하게 드러내라
- 그러나 페르소나 디테일에만 매몰되지 말고, 아래 [원래 역할 지침] 도 충실히 수행하라
═══════════════════════════════════════════

[원래 역할 지침]
{base_prompt}
"""
    return persona_prefix


def council_summary_md(council: dict[str, dict], stage: str = "") -> str:
    """Markdown summary of selected council for display + export.

    council: {"scout": persona_dict, "critic": persona_dict, ...}
    """
    if not council:
        return "_Council 미구성 (추상 Council 사용)_"

    lines = ["## Council 구성"]
    if stage:
        lines.append(f"_단계: {stage}_")
    lines.append("")
    role_order = ["scout", "critic", "director", "advisor"]
    for role in role_order:
        persona = council.get(role)
        if persona is None:
            continue
        label, _ = ROLE_DESCRIPTIONS.get(role, (role, ""))
        lines.append(f"### {label}")
        lines.append(f"- **{persona.get('이름', '?')}** ({persona.get('나이', '?')}세 {persona.get('성별', '?')}, {persona.get('시도', '?')} {persona.get('시군구', '')})")
        lines.append(f"- 직업: {persona.get('직업', '?')} / 학력: {persona.get('학력', '?')}")
        if persona.get("한줄_persona"):
            lines.append(f"- 요약: {persona['한줄_persona']}")
        lines.append("")
    return "\n".join(lines)
